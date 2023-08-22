# Name : Mohamed Amir Mohamed Abdullah .
# ID : 20211079 .
# Name : John Ayman Naim Aziz  .
# ID : 20210107 .
# Name : Ahmed yasser bastawessy .
# ID : 20210048 .
# The Application Number : 0 --> Excel application [ GUI ] by python .


from openpyxl import *
from tkinter import *

window = Tk()
window.geometry('1920x1080')
window.title('Excel')
window.config(bg='Green')
window.iconbitmap('Icon.ico')
bg = PhotoImage(file='Background.png')
back_ground = Label(window, image=bg).place(x=0, y=0)

global min_state
global max_state
global state
global total
global text
global state_input
global window_2


def get_state():
    text.delete(1.0, END)
    global state

    state = state_input.get()
    file = load_workbook('states.xlsx')

    if state in file.sheetnames:

        global total
        total = 0

        sheet = file[state]

        max_row = sheet.max_row

        for row in range(1, max_row + 1):
            cell = sheet.cell(row, 1)
            cell_2 = sheet.cell(row, 2)
            total += cell_2.value

            text.insert(1.0, "\n")
            text.insert(1.0, cell_2.value)
            text.insert(1.0, "   -->   ")
            text.insert(1.0, cell.value)

        text.insert(END, "\nTotal   -->  ")
        text.insert(END, total)

    else:

        text.insert(1.0, "The state is not exist ")


def min_max():
    text.delete(1.0, END)

    global min_state
    global max_state
    global state

    state = state_input.get()
    file = load_workbook('states.xlsx')

    if state in file.sheetnames:
        sheet = file[state]

        cell = sheet.cell(1, 2)
        max_pop = cell.value

        cell_2 = sheet.cell(1, 2)
        min_pop = cell_2.value

        max_row = sheet.max_row

        for row in range(1, max_row + 1):

            cell = sheet.cell(row, 2)
            val = cell.value

            if val >= max_pop:

                max_pop = val
                max_state = (sheet.cell(row, 1)).value

            else:
                pass

        for row in range(1, max_row + 1):

            cell_2 = sheet.cell(row, 2)
            val = cell_2.value

            if val <= min_pop:

                min_pop = val
                min_state = (sheet.cell(row, 1)).value

            else:
                pass

        text.insert(1.0, "The Max population is   -->   ")
        text.insert(2.0, max_pop)
        text.insert(3.0, "  -->   ")
        text.insert(4.0, max_state)

        text.insert(5.0, "\nThe Min population is   -->   ")
        text.insert(6.0, min_pop)
        text.insert(7.0, "  -->   ")
        text.insert(8.0, min_state)

    else:

        text.insert(1.0, "The state is not exist ")


def check():
    text.delete(1.0, END)

    while True:
        global state
        state = state_input.get()
        file = load_workbook('states.xlsx')

        if state in file.sheetnames:

            text.insert(1.0, "The state is exist ")
            break
        else:

            text.insert(1.0, "The state is not exist ")
            break


def close():
    window_2.destroy()
    window.destroy()


def start():
    global text
    global state_input
    global window_2

    window_2 = Tk()
    window_2.geometry('1920x1080')
    window_2.title('Excel')
    window_2.config(bg='Green')
    window_2.iconbitmap('Icon.ico')
    bg_2 = PhotoImage(file='Background.png')
    Label(window, image=bg_2).place(x=0, y=0)

    get_country = Button(window_2, width=67, height=3, text='Get the state and the total population', fg='Black',
                         font=('Times', 16, 'bold'), bd=4, bg='Yellow', activeforeground="Orange",
                         activebackground="Black", command=get_state)

    get_country.place(x=1091, y=200)

    check_country = Button(window_2, width=32, height=3, text='Check if the state exist or no', fg='Red',
                           font=('Times', 16, 'bold'), bd=4, bg='#000030', activeforeground="Black",
                           activebackground="Cyan", command=check)

    check_country.place(x=1090, y=100)

    get_min_max = Button(window_2, width=32, height=3, text='Get the min and max of state population ', fg='Yellow',
                         font=('Times', 16, 'bold'), bd=4, bg='Black', activeforeground="Black",
                         activebackground="Green", command=min_max)

    get_min_max.place(x=1509, y=100)

    text = Text(window_2, width=55, height=20, font=('Times', 20, 'bold'), fg='White', bg='#C04000', bd=4)
    text.place(x=40, y=250)

    state_input = Entry(window_2, width=29, bd=4, fg='White', font=('Times', 20, 'bold'), bg='#C04000', justify='left')

    state_input.place(x=400, y=158)

    input_state = Label(window_2, width=17, text="Enter the state here ... ", bd=4, fg='Black',
                        font=('Imprint MT Shadow', 21, 'bold'), bg='Green')

    input_state.place(x=40, y=150)

    exit_button = Button(window_2, width=64, height=3, text='Exit', fg='White',
                         font=('Times', 16, 'bold'), bd=4, bg='#36013F', command=close, activeforeground="#4A0404",
                         activebackground="White")

    exit_button.place(x=37, y=900)

    welcome = Label(window_2, text="Welcome to Excel application ", bd=4, fg='Black',
                    font=('Script MT Bold', 35, 'bold'), bg='Green')

    welcome.place(x=30, y=40)

    window.mainloop()


start_button = Button(window, width=36, height=3, text='Start', fg='Black',
                      font=('Times', 25, 'bold'), bd=5, bg='Green', command=start, activeforeground="#36013F",
                      activebackground="White")

start_button.place(x=100, y=600)


welcome_1 = Label(window, text="Welcome   to   Excel \nApplication  :) ", bd=4, fg='Black',
                  font=('Script MT Bold', 40, 'bold'), bg='Green')

welcome_1.place(x=750, y=50)

window.mainloop()
