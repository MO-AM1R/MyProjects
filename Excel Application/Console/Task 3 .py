# Name : Mohamed Amir Mohamed Abdullah .
# ID : 20211079 .
# Name : John Ayman Naim Aziz  .
# ID : 20210107 .
# Name : Ahmed yasser bastawessy .
# ID : 20210048 .
# The Application Number : 0 --> Excel application by python .


from openpyxl import *
global max_state
global min_state


def get_state():

    global total

    max_row = sheet.max_row

    for row in range(1, max_row + 1):

        cell = sheet.cell(row, 1)
        cell_2 = sheet.cell(row, 2)
        total += cell_2.value

        print(cell.value, " --> ", cell_2.value)
    print("\nThe total = ", total, "\n")


def min_max():

    global max_state
    global min_state

    cell = sheet.cell(1, 2)
    max_pop = cell.value

    cell_2 = sheet.cell(1, 2)
    min_pop = cell_2.value

    max_row = sheet.max_row

    for row in range(1, max_row + 1):

        cell = sheet.cell(row, 2)
        val = cell.value

        if val >= max_pop:

            max_pop = val
            max_state = (sheet.cell(row, 1)).value

        else:
            pass

    for row in range(1, max_row + 1):

        cell_2 = sheet.cell(row, 2)
        val = cell_2.value

        if val <= min_pop:

            min_pop = val
            min_state = (sheet.cell(row, 1)).value

        else:
            pass

    print("The Max population is --> ", max_pop, " and its state is --> ", max_state,
          "\nThe Min population is -->", min_pop, " and its state is --> ", min_state, "\n")


while True:

    state = input("Enter the state \nBut the first character uppercase \nUSA all characters uppercase\n"
                  "Without any space...")
    file = load_workbook('states.xlsx')

    if state in file.sheetnames:

        print("The state is exist ")
        sheet = file[state]
        check = int(input("Enter 1 - To load the country file\n\nEnter 2 - To get the Max population an the "
                          "Min population and their states in a particular state\n\nEnter 3 - To Exit\n--> "))

        if check == 1:

            total = 0

            get_state()

        elif check == 2:

            min_max()

        elif check == 3:

            exit()

        else:
            print("You input Invalid number ...")

    elif state.isdigit():

        print("Enter state ...")

    else:

        print("The state is not exist ...")
